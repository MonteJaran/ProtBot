"""
core/export_xlsx.py — the .xlsx half of ROADMAP.md item 5. openpyxl is a
declared dependency (pyproject.toml), so these tests import it directly
rather than skipping when it's missing — the same way tests/test_dataexport.py
doesn't skip on csv.
"""

import os
from datetime import datetime, timedelta

from openpyxl import load_workbook

from core import export_xlsx


_APPS = [
    {"id": 1, "name": "Discord", "category": "Social"},
    {"id": 2, "name": "Code", "category": "Productivity"},
]
_HISTORY = {
    1: [
        {"date": "2026-09-01", "total_sec": 3661},   # 1h 1m
        {"date": "2026-09-02", "total_sec": 59},      # 0m
    ],
    2: [
        {"date": "2026-09-01", "total_sec": 7200},   # 2h 0m
    ],
}


class TestBuildWorkbook:

    def test_header_row(self):
        ws = export_xlsx.build_workbook(_APPS, _HISTORY).active
        header = [cell.value for cell in ws[1]]
        assert header == [
            "App Name", "Category", "Date", "Total Time (sec)", "Total Time (h:m)",
        ]

    def test_header_is_styled(self):
        ws = export_xlsx.build_workbook(_APPS, _HISTORY).active
        for cell in ws[1]:
            assert cell.font.bold is True
            assert cell.fill.start_color.rgb == "00" + export_xlsx._HEADER_FILL_RGB

    def test_data_rows_follow_the_given_app_order(self):
        ws = export_xlsx.build_workbook(_APPS, _HISTORY).active
        rows = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
        assert rows == [
            ["Discord", "Social", "2026-09-01", 3661, "1h 1m"],
            ["Discord", "Social", "2026-09-02", 59, "0m"],
            ["Code", "Productivity", "2026-09-01", 7200, "2h 0m"],
        ]

    def test_sheet_is_named_usage(self):
        wb = export_xlsx.build_workbook(_APPS, _HISTORY)
        assert wb.active.title == "Usage"
        assert wb.sheetnames == ["Usage"]

    def test_header_row_stays_visible_when_scrolling(self):
        ws = export_xlsx.build_workbook(_APPS, _HISTORY).active
        assert ws.freeze_panes == "A2"

    def test_column_widths_are_set(self):
        ws = export_xlsx.build_workbook(_APPS, _HISTORY).active
        assert ws.column_dimensions["A"].width == 26
        assert ws.column_dimensions["E"].width == 16

    def test_an_app_with_no_history_is_skipped_not_an_error(self):
        apps = [{"id": 1, "name": "Idle App", "category": "Custom"}]
        ws = export_xlsx.build_workbook(apps, {}).active
        assert ws.max_row == 1   # header only

    def test_missing_history_key_for_an_app_is_also_fine(self):
        apps = [{"id": 99, "name": "Untracked Yet", "category": "Custom"}]
        ws = export_xlsx.build_workbook(apps, {1: [{"date": "x", "total_sec": 1}]}).active
        assert ws.max_row == 1

    def test_no_apps_at_all_produces_just_the_header(self):
        ws = export_xlsx.build_workbook([], {}).active
        assert ws.max_row == 1

    def test_none_apps_and_none_history_do_not_raise(self):
        ws = export_xlsx.build_workbook(None, None).active
        assert ws.max_row == 1

    def test_missing_name_or_category_falls_back_to_empty_string(self):
        apps = [{"id": 1}]
        history = {1: [{"date": "2026-09-01", "total_sec": 60}]}
        ws = export_xlsx.build_workbook(apps, history).active
        row = [c.value for c in ws[2]]
        assert row == ["", "", "2026-09-01", 60, "1m"]


class TestFmtHm:

    def test_under_an_hour(self):
        assert export_xlsx._fmt_hm(59) == "0m"
        assert export_xlsx._fmt_hm(60) == "1m"

    def test_over_an_hour(self):
        assert export_xlsx._fmt_hm(3661) == "1h 1m"

    def test_negative_or_missing_clamps_to_zero(self):
        assert export_xlsx._fmt_hm(-100) == "0m"
        assert export_xlsx._fmt_hm(None) == "0m"


class TestWriteExport:

    def test_writes_a_loadable_file_and_returns_the_path(self, tmp_path):
        path = str(tmp_path / "out.xlsx")
        result = export_xlsx.write_export(_APPS, _HISTORY, path)
        assert result == path
        assert os.path.exists(path)

        ws = load_workbook(path).active
        assert ws["A1"].value == "App Name"
        assert ws["A2"].value == "Discord"


class TestDefaultFilename:

    def test_format(self):
        name = export_xlsx.default_filename(datetime(2026, 9, 4, 13, 5, 9))
        assert name == "ProtBot_export_20260904_130509.xlsx"

    def test_ends_in_xlsx(self):
        assert export_xlsx.default_filename().endswith(".xlsx")

    def test_defaults_to_now_when_no_argument_given(self):
        before = datetime.now()
        name = export_xlsx.default_filename()
        after = datetime.now()
        stamp = name[len("ProtBot_export_"):-len(".xlsx")]
        parsed = datetime.strptime(stamp, "%Y%m%d_%H%M%S")
        # Second-precision formatting can floor `before`, so allow a 1s pad.
        assert before.replace(microsecond=0) - timedelta(seconds=1) <= parsed
        assert parsed <= after.replace(microsecond=0) + timedelta(seconds=1)
