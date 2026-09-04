"""
export_xlsx.py - .xlsx usage export (ROADMAP.md item 5, "PDF / Excel report
export" — the Excel half; PDF is still not started, see ROADMAP.md).

`ui/processes_page.py`'s CSV export already exists and works — this is the
same 30-day per-app usage history, as a formatted workbook instead of a flat
file: styled header row, sized columns, frozen header. Nothing here recomputes
what the CSV export already gets right; it is the same data, shaped for a
spreadsheet instead of a text file.

Separated from the UI the way core/dataexport.py is: building the workbook is
a plain function of its arguments — apps and their history in, an
openpyxl.Workbook out — so it is testable by reading the cells back, without
a database or a database fixture standing in for one.

openpyxl is imported lazily, inside the functions that need it, and a missing
install raises ExportUnavailable rather than ImportError — one exception type
the caller handles, the same shape core/monitor.py's plyer import degrades
with (AUDIT: "optional dependencies degrade gracefully"). Unlike plyer's
notifications, openpyxl is a declared runtime dependency (pyproject.toml) —
this is not "the feature is optional forever", it is "an environment missing
a declared dependency gets a clear error instead of an ImportError traceback
with no context."
"""

from datetime import datetime

_HEADERS = ["App Name", "Category", "Date", "Total Time (sec)", "Total Time (h:m)"]

# Matches ui/theme.py's BG3 — the export is meant to look like it came from
# this app, not like a bare openpyxl default.
_HEADER_FILL_RGB = "0F3460"


class ExportUnavailable(Exception):
    """openpyxl is not installed."""


def _fmt_hm(seconds: int) -> str:
    seconds = max(0, int(seconds or 0))
    hours, minutes = divmod(seconds // 60, 60)
    return f"{hours}h {minutes}m" if hours else f"{minutes}m"


def build_workbook(apps, history_by_app_id: dict):
    """
    A workbook with one row per (app, date) usage record, newest apps'
    history included in whatever order `apps` is given in.

    `apps` is core/database.py's get_all_tracked_apps() shape — needs "id",
    "name", "category". `history_by_app_id` maps each app's id to what
    get_usage_history(app_id) returns for it: a list of {"date",
    "total_sec"}. Both are fetched by the caller — this function does no
    I/O — so a missing or empty history for an app is silently skipped
    rather than treated as an error: an app with no usage yet is a normal
    state, not a broken export.

    Raises ExportUnavailable if openpyxl is not installed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ExportUnavailable("openpyxl is not installed") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Usage"

    ws.append(_HEADERS)
    header_fill = PatternFill(start_color=_HEADER_FILL_RGB,
                              end_color=_HEADER_FILL_RGB, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(_HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill

    history_by_app_id = history_by_app_id or {}
    for app in apps or ():
        history = history_by_app_id.get(app.get("id")) or []
        for entry in history:
            seconds = int(entry.get("total_sec", 0) or 0)
            ws.append([
                app.get("name", "") or "",
                app.get("category", "") or "",
                entry.get("date", "") or "",
                seconds,
                _fmt_hm(seconds),
            ])

    widths = [26, 16, 12, 18, 16]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"   # the header row stays visible when scrolling
    return wb


def write_export(apps, history_by_app_id: dict, path: str) -> str:
    """Build the workbook and save it to `path`. Returns `path`, for a
    caller that wants to report where the file went, the way
    ui/processes_page.py's export_csv already does."""
    build_workbook(apps, history_by_app_id).save(path)
    return path


def default_filename(now=None) -> str:
    """`ProtBot_export_<timestamp>.xlsx` — export_csv's own naming scheme,
    so the two sit next to each other sensibly on disk."""
    moment = now if isinstance(now, datetime) else datetime.now()
    return f"ProtBot_export_{moment.strftime('%Y%m%d_%H%M%S')}.xlsx"
